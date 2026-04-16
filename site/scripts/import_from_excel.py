"""One-shot seeder that reads the curated Excel workbook and emits MDX stubs.

The workbook lives at ~/Downloads/Learning-Engineering-Resources.xlsx and has three
sheets (Resources, ASU & CMU, IITSEC & AF). Each row is one curated resource with
columns: Resource (title), Source (venue), Type (format), Date (year),
Author/Presenter, Link, and optional Other Links. Sub-section headers are rows
where only the Resource column is filled; we use them to set the `cluster` tag and
to route items into org-anchored collections.

Routing rules, in order (first match wins):

    1. Sub-section header names a recurring community / venue with an org focus
       → emit an `org` record into `community/`, and route child rows by Type.
    2. Type is Video / Podcast / Keynote / Webinar / Talks / Conference-talk /
       Conference / Workshop / convenings / Talk-Video → `events/`.
    3. Type is Journal / Article / Paper / Report / Book / Post / Essay / Brochure
       / Fact Sheet → `reading-list/`.
    4. Type is Tool / Template / Toolkit / Platform / Resources → route between
       `tools/`, `practice/`, or `methods` subgroup of practice by name keywords.
    5. Untyped rows → emit with missingProvenance=true into the best-guess
       collection based on sub-section (default: community/ for org sub-sections,
       reading-list/ otherwise).

Run from the repo root:

    source venv/bin/activate
    python3 site/scripts/import_from_excel.py
"""

from __future__ import annotations

import re
from datetime import date as _date
from pathlib import Path

try:
    import openpyxl
except ImportError as exc:  # keep the error explicit so users activate venv first
    raise SystemExit("openpyxl missing. Run: source venv/bin/activate") from exc

ROOT = Path(__file__).resolve().parents[2]
SITE = ROOT / "site"
CONTENT = SITE / "src" / "content"
XLSX = Path.home() / "Downloads" / "Learning-Engineering-Resources.xlsx"

# Sub-section headers that represent an org / community / recurring venue.
# Each maps to: (community_item_title, cluster_tag, child_default_section).
ORG_HEADERS: dict[str, tuple[str, str, str]] = {
    "LE @ CMU": (
        "Carnegie Mellon Open Learning Initiative & Simon Initiative",
        "CMU",
        "community",
    ),
    "Learning Engineering Institute, ASU": (
        "Learning Engineering Institute (ASU)",
        "ASU LEI",
        "community",
    ),
    "Air Force Enterprise Learning Engineering Center of Excellence": (
        "Air Force Enterprise Learning Engineering Center of Excellence",
        "AF ELE CoE",
        "community",
    ),
    "I/ITSEC Conference": (
        "I/ITSEC — Interservice/Industry Training, Simulation & Education Conference",
        "I/ITSEC",
        "events",
    ),
}

# Curatorial cluster tags for non-org sub-sections (applied to child rows).
NON_ORG_CLUSTERS: dict[str, str] = {
    "Background": "Background",
    "Books": "",
    "Bror Saxberg - He's been talking about LE for years": "Bror Saxberg",
    "Resources: Start with this list.  I realized these may be more from an ICICLE perspective.  Lots more are the list in blue.": "ICICLE resources",
    "And More Resources": "",
    "Wikipedia": "Wikipedia",
    "LE page": "Wikipedia",
}


# Type → (collection, format). Case-insensitive match on the *entire* Type string.
TYPE_MAP: dict[str, tuple[str, str]] = {
    # reading-list formats
    "journal":       ("reading-list", "paper"),
    "article":       ("reading-list", "article"),
    "paper":         ("reading-list", "paper"),
    "preprint":      ("reading-list", "paper"),
    "book":          ("reading-list", "book"),
    "post":          ("reading-list", "post"),
    "essay":         ("reading-list", "essay"),
    "report":        ("reading-list", "report"),
    "brochure":      ("reading-list", "report"),
    "fact sheet":    ("reading-list", "report"),
    "news":          ("reading-list", "article"),

    # events formats
    "video":              ("events", "video"),
    "podcast":            ("events", "podcast"),
    "keynote":            ("events", "keynote"),
    "webinar":            ("events", "webinar"),
    "talk":               ("events", "keynote"),
    "talks":              ("events", "keynote"),
    "talk - video":       ("events", "keynote"),
    "conference":         ("events", "conference-talk"),
    "conference-talk":    ("events", "conference-talk"),
    "workshop":           ("events", "workshop"),
    "convening":          ("events", "convening"),
    "convenings":         ("events", "series"),
    "series":             ("events", "series"),

    # tools formats
    "tool":       ("tools", "tool"),
    "platform":   ("tools", "platform"),
    "template":   ("tools", "template"),
    "toolkit":    ("tools", "toolkit"),
    "resources":  ("tools", "platform"),
}

# Name-keyword overrides pulling tools-tagged rows into the practice collection
# when the item is clearly about the LE process or an analysis method.
PRACTICE_KEYWORDS = (
    "process", "five whys", "fishbone", "task analysis", "checklist",
    "design principles", "maturity model", "evidence decision",
    "case guide", "implementation",
)


def slugify(text: str) -> str:
    """Convert title text to a filesystem-safe slug, clipped to 70 chars."""
    s = re.sub(r"[^a-z0-9]+", "-", text.lower()).strip("-")
    return s[:70] or "untitled"


def yaml_escape(s: str) -> str:
    """Quote a string value for YAML frontmatter output."""
    return '"' + s.replace("\\", "\\\\").replace('"', '\\"') + '"'


def write_mdx(
    collection: str,
    slug: str,
    frontmatter: dict,
    body: str,
) -> bool:
    """Write a single MDX file unless a file with the same slug already exists."""
    out_dir = CONTENT / collection
    out_dir.mkdir(parents=True, exist_ok=True)
    path = out_dir / f"{slug}.mdx"
    if path.exists():
        return False

    lines: list[str] = ["---"]

    def put(key: str, value):
        if value is None or value == "" or value == []:
            return
        if isinstance(value, bool):
            if value:
                lines.append(f"{key}: true")
            return
        if isinstance(value, (int, float)):
            lines.append(f"{key}: {value}")
            return
        if isinstance(value, list):
            lines.append(f"{key}:")
            for item in value:
                lines.append(f"  - {yaml_escape(str(item))}")
            return
        lines.append(f"{key}: {yaml_escape(str(value))}")

    put("title", frontmatter["title"])
    put("format", frontmatter["format"])
    put("venue", frontmatter.get("venue"))
    put("authors", frontmatter.get("authors"))
    if frontmatter.get("year"):
        put("year", int(frontmatter["year"]))
    put("url", frontmatter.get("url"))
    put("otherUrls", frontmatter.get("otherUrls", []))
    put("cluster", frontmatter.get("cluster"))
    put("topics", frontmatter.get("topics", []))
    put("tags", frontmatter.get("tags", []))
    put("featured", frontmatter.get("featured", False))
    put("missingProvenance", frontmatter.get("missingProvenance", False))

    prov = frontmatter["provenance"]
    lines.append("provenance:")
    lines.append(f"  dataset: {yaml_escape(prov['dataset'])}")
    if prov.get("ref"):
        lines.append(f"  ref: {yaml_escape(prov['ref'])}")
    if prov.get("sheet"):
        lines.append(f"  sheet: {yaml_escape(prov['sheet'])}")
    if prov.get("sectionHeader"):
        lines.append(f"  sectionHeader: {yaml_escape(prov['sectionHeader'])}")

    lines.append("---")
    lines.append("")
    lines.append(body.strip() or "_No description yet._")
    lines.append("")

    path.write_text("\n".join(lines), encoding="utf-8")
    return True


def classify_row(
    resource: str,
    type_: str | None,
    cluster: str | None,
    org_default: str | None,
) -> tuple[str, str]:
    """Return (collection, format) for a data row. Handles overrides and fallbacks."""
    name_lc = resource.lower()
    type_key = (type_ or "").strip().lower()

    # Name-keyword override: practice-style items always go to practice/
    if any(k in name_lc for k in PRACTICE_KEYWORDS):
        # Pick a precise format when possible
        if any(k in name_lc for k in ("five whys", "fishbone", "task analysis")):
            return "practice", "method"
        if "checklist" in name_lc or "maturity" in name_lc:
            return "practice", "framework"
        if "case guide" in name_lc or "design principles" in name_lc:
            return "practice", "framework"
        return "practice", "framework"

    # Direct type lookup
    if type_key in TYPE_MAP:
        return TYPE_MAP[type_key]

    # Bare rows (no type): use the org default if we have one, else reading-list
    if not type_key:
        if org_default == "community":
            return "community", "org"
        if org_default == "events":
            return "events", "conference-talk"
        return "reading-list", "article"

    # Unknown type: conservative fallback
    return "reading-list", "article"


def parse_year(date_value) -> int | None:
    """Extract a 4-digit year from the Excel Date cell (int / str / datetime)."""
    if isinstance(date_value, (int, float)):
        y = int(date_value)
        if 1900 < y < 2100:
            return y
    if isinstance(date_value, _date):
        return date_value.year
    s = str(date_value or "")
    m = re.search(r"\b(19|20)\d{2}\b", s)
    return int(m.group(0)) if m else None


def is_header_row(values: list) -> bool:
    """Detect rows that are purely section headers — no index, resource filled, no metadata.

    Sheets 2 and 3 of the workbook use column 0 as an item index for data rows.
    A row with any index value (even whitespace like '   ') is a data row even
    when its metadata columns are empty — this is how we preserve the bare AF
    ELE CoE titles that have no source/link/date yet.
    """
    idx, resource, src, tp, date, author, link, *_ = (values + [None] * 8)[:8]
    if idx is not None:
        return False
    return bool(resource) and not any([src, tp, date, author, link])


def walk_sheet(ws, sheet_name: str, stats: dict) -> None:
    """Process one Excel sheet row-by-row, emitting MDX per data row."""
    current_header = ""
    for row_idx, row in enumerate(ws.iter_rows(values_only=True), 1):
        if row_idx == 1:
            continue  # column names
        values = list(row)
        _, resource, src, tp, date_v, author, link, *rest = (values + [None] * 8)[:8]
        if not any([resource, src, tp, date_v, author, link]):
            continue

        if is_header_row(values):
            current_header = str(resource).strip()
            # If the header itself is an org, emit the org record up front.
            if current_header in ORG_HEADERS:
                title, cluster_tag, _default = ORG_HEADERS[current_header]
                slug = slugify(title)
                write_mdx(
                    "community",
                    slug,
                    {
                        "title": title,
                        "format": "org",
                        "venue": "",
                        "cluster": cluster_tag,
                        "missingProvenance": False,
                        "provenance": {
                            "dataset": "LE Resources Excel v1",
                            "sheet": sheet_name,
                            "ref": f"{sheet_name}!row{row_idx}",
                            "sectionHeader": current_header,
                        },
                    },
                    f"Curatorial cluster header from the LE Resources workbook, sheet '{sheet_name}'.",
                )
                stats["community"] = stats.get("community", 0) + 1
            continue

        # Data row — route it.
        resource = str(resource).strip()
        if current_header in ORG_HEADERS:
            _, cluster_tag, default_collection = ORG_HEADERS[current_header]
        else:
            cluster_tag = NON_ORG_CLUSTERS.get(current_header, "")
            default_collection = None

        collection, fmt = classify_row(resource, tp, cluster_tag, default_collection)

        other_urls = [u for u in rest if isinstance(u, str) and u.strip().startswith(("http://", "https://"))]

        slug = slugify(resource)
        missing = not any([src, tp, link, author, date_v])

        fm = {
            "title": resource,
            "format": fmt,
            "venue": str(src).strip() if src else None,
            "authors": str(author).strip() if author else None,
            "year": parse_year(date_v),
            "url": str(link).strip() if link else None,
            "otherUrls": other_urls,
            "cluster": cluster_tag or None,
            "missingProvenance": missing,
            "provenance": {
                "dataset": "LE Resources Excel v1",
                "sheet": sheet_name,
                "ref": f"{sheet_name}!row{row_idx}",
                "sectionHeader": current_header or None,
            },
        }
        body_parts: list[str] = []
        if fm["venue"]:
            body_parts.append(f"Published in **{fm['venue']}**.")
        if fm["authors"]:
            body_parts.append(f"By {fm['authors']}.")
        if not body_parts:
            body_parts.append("_Blurb pending._")
        body = " ".join(body_parts)

        if write_mdx(collection, slug, fm, body):
            stats[collection] = stats.get(collection, 0) + 1


def main() -> None:
    """Load the workbook and seed all six collections."""
    if not XLSX.exists():
        raise SystemExit(f"Excel file not found at {XLSX}")

    wb = openpyxl.load_workbook(XLSX, data_only=True)
    stats: dict[str, int] = {}
    for name in wb.sheetnames:
        walk_sheet(wb[name], name, stats)

    print("Wrote MDX files:")
    for section in ("practice", "tools", "reading-list", "events", "community"):
        print(f"  {section:13s} +{stats.get(section, 0)}")
    print(f"  total         {sum(stats.values())}")


if __name__ == "__main__":
    main()
