#!/usr/bin/env python3
"""Ingest curated LE resources from an Excel workbook into non_paper_resources.jsonl.

Reads up to 3 tabs, extracts rows with title + URL, deduplicates within the workbook
and against the existing corpus, then appends new records as GL (grey_literature)
entries with heuristically assigned topic codes.
"""

from __future__ import annotations

import argparse
import json
import re
import sys
from pathlib import Path
from typing import Optional

ROOT = Path(__file__).resolve().parents[1]
CORPUS_DIR = ROOT / "corpus"
NON_PAPER_PATH = CORPUS_DIR / "non_paper_resources.jsonl"

# ---------------------------------------------------------------------------
# Topic heuristics: ordered list of (keyword_pattern, primary_topic, secondary_topics)
# Applied to title + source + section text (lowercased).
# ---------------------------------------------------------------------------
TOPIC_RULES: list[tuple[str, str, list[str]]] = [
    (r"iitsec|i/itsec|\bmilitary\b|defense|dod|\barmy\b|air force|training system|workforce develop", "T10", ["T13", "T00"]),
    (r"simulat|\bxr\b|\bvr\b|virtual reality|augmented reality|serious game|experiential learn", "T08", ["T10", "T00"]),
    (r"ai tutor|intelligent tutor|cognitive tutor|its system|adaptive learning|knowledge trac|bayesian student", "T06", ["T07", "T00"]),
    (r"\bllm\b|large language model|generative ai|chatgpt|gpt-\d|foundation model", "T07", ["T06", "T00"]),
    (r"learning analytic|data mining|\bxapi\b|\blrs\b|learning record|psychometric|assessment data", "T04", ["T03", "T00"]),
    (r"knowledge graph|ontolog|competency framework|skill taxonom|knowledge represent", "T05", ["T04", "T00"]),
    (r"learning infrastructure|\blms\b|platform interoperab|open standard|tin can", "T11", ["T04", "T00"]),
    (r"instructional design|curriculum|\baddie\b|backward design|learning objective", "T12", ["T03", "T00"]),
    (r"cognitive task analysis|expert knowledge elicit|tacit knowledge|\bsme\b elicit", "T09", ["T12", "T00"]),
    (r"ethics|equity|fairness|\bbias\b|privacy|responsible ai|access gap", "T14", ["T07", "T00"]),
    (r"kirkpatrick|\brct\b|randomized.controlled|reproducib|replication|what counts as evidence", "T15", ["T17", "T00"]),
    (r"ieee icicle|icicle|body of knowledge|\bbok\b|professional community|credenti", "T16", ["T00"]),
    (r"research method|design.based research|open science|field develop", "T17", ["T15", "T00"]),
    (r"human factor|systems engineering|human.computer interaction|\bhsi\b|sociotechnical", "T02", ["T03", "T00"]),
    (r"cognitive load|spaced practice|retrieval practice|how people learn", "T01", ["T00"]),
    (r"learning engineering process|\ble process\b|iterative design|evidence.based design|rapid prototyp", "T03", ["T00"]),
]
DEFAULT_TOPIC = "T00"
DEFAULT_SECONDARY: list[str] = []

# Map broad content types from the spreadsheet → our content_type codes
TYPE_MAP: dict[str, str] = {
    "journal": "GL",
    "journal publication": "GL",
    "conference": "GL",
    "conference paper": "GL",
    "conference proceeding": "GL",
    "conference proceedings": "GL",
    "conference session": "GL",
    "paper": "GL",
    "book": "GL",
    "book chapter": "GL",
    "point paper. panel": "GL",
    "post": "GL",
    "post ": "GL",
    "blog post": "GL",
    "blog posts": "GL",
    "article": "GL",
    " journal": "GL",
    "report": "GL",
    "podcast": "GL",
    "video": "GL",
    "webinar": "GL",
    "talk": "GL",
    "keynote": "GL",
    "talk - video": "GL",
    "video/podcast": "GL",
    "poster": "GL",
    "resources": "GL",
    "brochure": "GL",
    "convenings": "CE",
    "talks": "GL",
    "videos on what is learning engineering based on toolkit": "GL",
    "journal article": "GL",
}


def normalize_url(url: str) -> str:
    """Strip whitespace and trailing slashes for comparison."""
    return re.sub(r"\s+", "", (url or "")).rstrip("/").lower()


def _clean_title(raw: str) -> str:
    return re.sub(r"\s+", " ", (raw or "").replace("\n", " ").replace("\xa0", " ")).strip()


def assign_topics(title: str, source: str, section: str) -> tuple[str, list[str]]:
    """Return (primary_topic, secondary_topics) based on heuristic keyword scan."""
    haystack = " ".join([title, source, section]).lower()
    for pattern, primary, secondary in TOPIC_RULES:
        if re.search(pattern, haystack):
            return primary, secondary
    return DEFAULT_TOPIC, DEFAULT_SECONDARY


def load_existing(path: Path) -> list[dict]:
    """Load existing JSONL records."""
    if not path.exists():
        return []
    return [json.loads(line) for line in path.read_text(encoding="utf-8").splitlines() if line.strip()]


def next_gl_id(existing: list[dict]) -> int:
    """Return the next integer suffix for LE-GL-NNN IDs."""
    max_n = 0
    for rec in existing:
        rid = rec.get("resource_id", "")
        m = re.match(r"LE-GL-(\d+)", rid)
        if m:
            max_n = max(max_n, int(m.group(1)))
    return max_n + 1


def parse_excel(xlsx_path: Path) -> list[dict]:
    """Parse all three tabs and return a flat list of candidate rows."""
    try:
        import openpyxl  # noqa: PLC0415
    except ImportError:
        sys.exit("openpyxl is required: pip install openpyxl")

    wb = openpyxl.load_workbook(str(xlsx_path), data_only=True)
    rows: list[dict] = []

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        current_section: Optional[str] = None

        for r in range(2, ws.max_row + 1):
            title_raw = ws.cell(row=r, column=2).value
            source = ws.cell(row=r, column=3).value
            rtype = ws.cell(row=r, column=4).value
            date_val = ws.cell(row=r, column=5).value
            author = ws.cell(row=r, column=6).value
            link = ws.cell(row=r, column=7).value
            other_link = ws.cell(row=r, column=8).value

            if not title_raw:
                continue

            title = _clean_title(str(title_raw))

            # Detect section-header rows: title present but no source/type/link
            if not source and not rtype and not link:
                current_section = title
                continue

            # Resolve URL — prefer col7, fall back to col8 if it looks like a URL
            url = str(link).strip() if link else ""
            if not url.startswith("http") and other_link and str(other_link).startswith("http"):
                url = str(other_link).strip()
            url = re.sub(r"\s+", "", url)  # strip internal spaces

            # Skip rows with no meaningful data beyond a title
            if not url and not source and not rtype:
                continue

            year: Optional[int] = None
            if date_val:
                m = re.match(r"(\d{4})", str(date_val))
                if m:
                    year = int(m.group(1))

            rows.append({
                "sheet": sheet_name,
                "section": current_section or "",
                "title": title,
                "source": str(source).strip() if source else "",
                "type": str(rtype).strip() if rtype else "",
                "year": year,
                "author": str(author).strip() if author else "",
                "url": url,
            })

    return rows


def build_records(
    candidates: list[dict],
    existing: list[dict],
) -> tuple[list[dict], int, int]:
    """Deduplicate candidates and return new JSONL records to append.

    Returns (new_records, n_skipped_existing, n_skipped_internal_dup).
    """
    existing_urls: set[str] = {normalize_url(r.get("url", "")) for r in existing}
    existing_names_norm: set[str] = {r.get("name", "").lower().strip() for r in existing}

    seen_urls: set[str] = set()
    seen_titles: set[str] = set()
    skipped_existing = 0
    skipped_dup = 0
    new_records: list[dict] = []
    counter = next_gl_id(existing)

    for row in candidates:
        url_norm = normalize_url(row["url"])
        title_norm = row["title"].lower()

        # Dedup against existing corpus
        if url_norm and url_norm in existing_urls:
            skipped_existing += 1
            continue
        if title_norm in existing_names_norm:
            skipped_existing += 1
            continue

        # Dedup within the Excel itself
        if url_norm and url_norm in seen_urls:
            skipped_dup += 1
            continue
        if title_norm in seen_titles:
            skipped_dup += 1
            continue

        if url_norm:
            seen_urls.add(url_norm)
        seen_titles.add(title_norm)

        primary, secondary = assign_topics(row["title"], row["source"], row["section"])

        content_type = TYPE_MAP.get(row["type"].lower(), "GL") if row["type"] else "GL"

        description_parts = []
        if row["source"]:
            description_parts.append(f"Source: {row['source']}.")
        if row["author"]:
            description_parts.append(f"Author(s): {row['author']}.")
        if row["type"]:
            description_parts.append(f"Type: {row['type']}.")
        if row["year"]:
            description_parts.append(f"Year: {row['year']}.")

        record = {
            "resource_id": f"LE-GL-{counter:03d}",
            "record_type": "non_paper_resource",
            "status": "APPROVED",
            "content_type": content_type,
            "name": row["title"],
            "affiliation_or_venue": row["source"],
            "url": row["url"],
            "primary_topic": primary,
            "secondary_topics": secondary,
            "description": " ".join(description_parts),
            "notes": f"Ingested from Excel workbook tab '{row['sheet']}', section '{row['section']}'.",
            "year": row["year"],
            "author": row["author"],
            "selection_tier": "",
            "selection_source": "excel_workbook",
            "ieee_icicle_listed": "",
            "reviewed_by": "",
            "date_reviewed": "",
        }
        new_records.append(record)
        counter += 1

    return new_records, skipped_existing, skipped_dup


def main() -> None:
    """Parse CLI args and run the ingestion."""
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("xlsx", type=Path, help="Path to the Excel workbook (.xlsx)")
    parser.add_argument("--dry-run", action="store_true", help="Print results without writing")
    args = parser.parse_args()

    if not args.xlsx.exists():
        sys.exit(f"File not found: {args.xlsx}")

    print(f"Parsing {args.xlsx} …")
    candidates = parse_excel(args.xlsx)
    print(f"  Extracted {len(candidates)} candidate rows across all sheets")

    existing = load_existing(NON_PAPER_PATH)
    print(f"  Existing corpus records: {len(existing)}")

    new_records, skipped_existing, skipped_dup = build_records(candidates, existing)

    print(f"\nDeduplication summary:")
    print(f"  Skipped (already in corpus): {skipped_existing}")
    print(f"  Skipped (internal Excel duplicates): {skipped_dup}")
    print(f"  New records to add: {len(new_records)}")

    if not new_records:
        print("Nothing to add.")
        return

    print("\nNew records:")
    for r in new_records:
        print(f"  {r['resource_id']} [{r['content_type']}|{r['primary_topic']}] {r['name'][:70]}")
        if r["url"]:
            print(f"    {r['url'][:80]}")

    if args.dry_run:
        print("\n[dry-run] Not writing to disk.")
        return

    with NON_PAPER_PATH.open("a", encoding="utf-8") as fh:
        for record in new_records:
            fh.write(json.dumps(record, ensure_ascii=False) + "\n")

    print(f"\nAppended {len(new_records)} records to {NON_PAPER_PATH}")
    print("Next step: run  python scripts/build_dataset.py  to regenerate data/*.json")


if __name__ == "__main__":
    main()
